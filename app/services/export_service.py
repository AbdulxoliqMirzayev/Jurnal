from __future__ import annotations

from datetime import datetime
from pathlib import Path

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database.models import DepositTransaction, EmotionStats, ExportHistory, TradeJournal, User, UserProfile
from app.services.analytics_service import ai_action_plan, emotion_counts, equity_curve, instrument_pnl, performance_score, risk_buckets
from app.services.stats_service import JournalStats, StatsService, stats_advice
from app.utils.dates import parse_period_days

EXPORT_DIR = Path("exports")


class ExportService:
    def __init__(self, session: AsyncSession) -> None:
        self.session = session

    async def pdf(self, user: User, stats: JournalStats, profile: UserProfile | None, period: str, ai_notes: str | None = None) -> str:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        path = EXPORT_DIR / f"iron_trade_{user.telegram_id}_{period}_{_stamp()}.pdf"
        doc = SimpleDocTemplate(str(path), pagesize=A4, rightMargin=1.1 * cm, leftMargin=1.1 * cm, topMargin=1.0 * cm)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "IronTitle",
            parent=styles["Title"],
            textColor=colors.HexColor("#F7C948"),
            fontSize=22,
            leading=26,
        )
        section_style = ParagraphStyle(
            "IronSection",
            parent=styles["Heading2"],
            textColor=colors.HexColor("#111827"),
            fontSize=13,
        )
        dark_style = ParagraphStyle(
            "IronDark",
            parent=styles["Normal"],
            textColor=colors.white,
            leading=15,
        )
        normal = styles["Normal"]
        score = performance_score(stats)
        story = [
            _hero_table(
                title_style,
                dark_style,
                user.full_name or user.username or str(user.telegram_id),
                stats.period_label,
                user.trading_type or "-",
                score,
            ),
            Spacer(1, 10),
            _summary_cards(stats, profile),
            Spacer(1, 12),
        ]
        story.extend(
            [
                Paragraph("Visual Analytics", section_style),
                _chart_grid(stats),
                Spacer(1, 14),
                Paragraph("Iron AI Action Plan", section_style),
                Paragraph((ai_notes or ai_action_plan(stats)).replace("\n", "<br/>"), normal),
                Spacer(1, 10),
                Paragraph("AI Notes", section_style),
                Paragraph(stats_advice(stats).replace("\n", "<br/>"), normal),
                Spacer(1, 12),
            ]
        )
        trade_rows = [["Date", "Instrument", "Trades", "Risk %", "Net", "Emotion", "Mistakes"]]
        for entry in stats.entries[:80]:
            trade_rows.append(
                [
                    entry.date.isoformat(),
                    entry.instrument or entry.pair or entry.coin_symbol or "-",
                    str(entry.trade_count),
                    "" if entry.risk_percent is None else f"{entry.risk_percent:g}",
                    f"{entry.net_result:+.2f}",
                    entry.emotion or "-",
                    ", ".join(entry.mistakes or [])[:38],
                ]
            )
        trade_table = Table(trade_rows, repeatRows=1)
        trade_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#94A3B8")),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                ]
            )
        )
        story.extend([Paragraph("Trade List", section_style), trade_table, Spacer(1, 12)])
        story.append(Paragraph("Disclaimer: This is not financial advice. The report is based only on your journal data.", normal))
        doc.build(story)
        await self._record(user, "pdf", period, str(path))
        return str(path)

    async def excel(self, user: User, stats: JournalStats, profile: UserProfile | None, period: str, ai_notes: str | None = None) -> str:
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        path = EXPORT_DIR / f"iron_trade_{user.telegram_id}_{period}_{_stamp()}.xlsx"
        wb = Workbook()
        header_fill = PatternFill("solid", fgColor="111827")
        header_font = Font(bold=True, color="FFFFFF")

        summary = wb.active
        summary.title = "Summary"
        summary.append(["Metric", "Value"])
        for row in [
            ("User", user.full_name or user.username or str(user.telegram_id)),
            ("Trading type", user.trading_type or "-"),
            ("Period", stats.period_label),
            ("Deposit", profile.deposit_current if profile else 0),
            ("Total trades", stats.total_trades),
            ("Win rate", stats.win_rate),
            ("Net PnL", stats.net_pnl),
            ("Best instrument", stats.best_instrument or "-"),
            ("Worst instrument", stats.worst_instrument or "-"),
            ("Best session", stats.best_session or "-"),
            ("Worst session", stats.worst_session or "-"),
        ]:
            summary.append(row)
        _style_header(summary, header_fill, header_font)

        trades = wb.create_sheet("Trades")
        trades.append(["Date", "Market", "Pair/Coin", "Session", "Trade Count", "Risk %", "Profit", "Loss", "Net Result", "Emotion", "Mistakes", "AI Score", "Notes"])
        _style_header(trades, header_fill, header_font)
        for entry in stats.entries:
            trades.append(
                [
                    entry.date.isoformat(),
                    entry.trading_type,
                    entry.instrument or entry.pair or entry.coin_symbol,
                    entry.session,
                    entry.trade_count,
                    entry.risk_percent,
                    entry.profit_amount,
                    entry.loss_amount,
                    entry.net_result,
                    entry.emotion,
                    ", ".join(entry.mistakes or []),
                    entry.ai_score,
                    entry.ai_summary,
                ]
            )
        trades.auto_filter.ref = trades.dimensions

        instruments = wb.create_sheet("Instruments")
        instruments.append(["Instrument", "Total", "Wins", "Losses", "Win Rate", "Net", "Avg Risk", "Best", "Worst", "Mistake", "Emotion"])
        _style_header(instruments, header_fill, header_font)
        for row in await StatsService(self.session).instrument_rows(user, parse_period_days(period)):
            instruments.append([row["instrument"], row["total"], row["wins"], row["losses"], row["win_rate"], row["net"], row["avg_risk"], row["best_result"], row["worst_result"], row["mistake"], row["emotion"]])

        emotions = wb.create_sheet("Emotions")
        emotions.append(["Emotion", "Total Trades", "Wins", "Losses", "Net Result"])
        _style_header(emotions, header_fill, header_font)
        emotion_rows = await self.session.scalars(select(EmotionStats).where(EmotionStats.user_id == user.id))
        for emotion in emotion_rows:
            emotions.append([emotion.emotion, emotion.total_trades, emotion.win_count, emotion.loss_count, emotion.net_result])

        deposits = wb.create_sheet("Deposit History")
        deposits.append(["Date", "Type", "Amount", "Note"])
        _style_header(deposits, header_fill, header_font)
        deposit_rows = await self.session.scalars(select(DepositTransaction).where(DepositTransaction.user_id == user.id).order_by(DepositTransaction.created_at))
        for item in deposit_rows:
            deposits.append([item.created_at.strftime("%Y-%m-%d %H:%M"), item.type, item.amount, item.note])

        notes = wb.create_sheet("AI Notes")
        notes.append(["AI Summary"])
        notes.append([ai_notes or stats_advice(stats)])
        notes.append([])
        notes.append(["Iron AI Action Plan"])
        for line in ai_action_plan(stats).splitlines():
            notes.append([line])
        _style_header(notes, header_fill, header_font)

        charts = wb.create_sheet("Charts")
        charts.append(["Equity Date", "Cumulative PnL"])
        for point_date, value in equity_curve(stats.entries):
            charts.append([point_date.isoformat(), value])
        charts.append([])
        instrument_start = charts.max_row + 1
        charts.append(["Instrument", "Net PnL"])
        for name, value in instrument_pnl(stats.entries):
            charts.append([name, value])
        emotion_start = charts.max_row + 2
        charts.cell(emotion_start, 1, "Emotion")
        charts.cell(emotion_start, 2, "Count")
        for offset, (name, value) in enumerate(emotion_counts(stats.entries), 1):
            charts.cell(emotion_start + offset, 1, name)
            charts.cell(emotion_start + offset, 2, value)
        _style_header(charts, header_fill, header_font)
        _add_excel_charts(charts, instrument_start, emotion_start)
        for sheet in wb.worksheets:
            for col in range(1, sheet.max_column + 1):
                sheet.column_dimensions[sheet.cell(1, col).column_letter].width = 18
        wb.save(path)
        await self._record(user, "excel", period, str(path))
        return str(path)

    async def _record(self, user: User, export_type: str, period: str, path: str) -> None:
        self.session.add(ExportHistory(user_id=user.id, export_type=export_type, period=period, file_path=path))
        await self.session.flush()


def _style_header(sheet, fill, font) -> None:
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font


def _stamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _hero_table(title_style, dark_style, username: str, period: str, trading_type: str, score: int):
    from reportlab.lib import colors
    from reportlab.platypus import Paragraph, Table, TableStyle

    title = Paragraph("Iron Trade Journal Report", title_style)
    subtitle = Paragraph(
        f"AI-powered discipline report<br/>User: <b>{username}</b><br/>Period: <b>{period}</b> | Market: <b>{trading_type}</b>",
        dark_style,
    )
    score_text = Paragraph(f"<b>{score}/100</b><br/>Discipline score", dark_style)
    table = Table([[title, score_text], [subtitle, ""]], colWidths=[390, 110], rowHeights=[42, 54])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#0F172A")),
                ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#F7C948")),
                ("SPAN", (1, 0), (1, 1)),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (1, 0), (1, 1), "CENTER"),
                ("LEFTPADDING", (0, 0), (-1, -1), 14),
                ("RIGHTPADDING", (0, 0), (-1, -1), 14),
                ("TOPPADDING", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ]
        )
    )
    return table


def _summary_cards(stats: JournalStats, profile: UserProfile | None):
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle

    rows = [
        ["Total Trades", "Win Rate", "Net PnL", "Avg Risk"],
        [
            str(stats.total_trades),
            f"{stats.win_rate:.1f}%",
            f"{stats.net_pnl:+.2f}$",
            f"{stats.avg_risk:.1f}%",
        ],
        ["Best Instrument", "Worst Instrument", "Best Session", "Deposit"],
        [
            stats.best_instrument or "-",
            stats.worst_instrument or "-",
            stats.best_session or "-",
            f"{profile.deposit_current if profile else 0:.2f}$",
        ],
    ]
    table = Table(rows, colWidths=[125, 125, 125, 125])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
                ("BACKGROUND", (0, 2), (-1, 2), colors.HexColor("#111827")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("TEXTCOLOR", (0, 2), (-1, 2), colors.white),
                ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#F8FAFC")),
                ("BACKGROUND", (0, 3), (-1, 3), colors.HexColor("#F8FAFC")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, 2), (-1, 2), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
                ("PADDING", (0, 0), (-1, -1), 7),
            ]
        )
    )
    return table


def _chart_grid(stats: JournalStats):
    from reportlab.platypus import Table

    return Table(
        [
            [_equity_chart(stats.entries), _instrument_chart(stats.entries)],
            [_emotion_chart(stats.entries), _risk_chart(stats.entries)],
        ],
        colWidths=[250, 250],
        rowHeights=[150, 150],
    )


def _equity_chart(entries: list[TradeJournal]):
    from reportlab.graphics.charts.lineplots import LinePlot
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.lib import colors

    points = equity_curve(entries)
    drawing = Drawing(240, 140)
    drawing.add(String(8, 124, "Equity Curve", fontSize=9, fillColor=colors.HexColor("#111827")))
    if len(points) < 2:
        drawing.add(String(18, 62, "Not enough data", fontSize=8))
        return drawing
    chart = LinePlot()
    chart.x = 20
    chart.y = 20
    chart.width = 200
    chart.height = 90
    chart.data = [[(idx + 1, value) for idx, (_, value) in enumerate(points)]]
    chart.lines[0].strokeColor = colors.HexColor("#2563EB")
    chart.lines[0].strokeWidth = 2
    chart.xValueAxis.valueMin = 1
    chart.xValueAxis.valueMax = len(points)
    values = [value for _, value in points]
    chart.yValueAxis.valueMin = min(values + [0])
    chart.yValueAxis.valueMax = max(values + [0]) or 1
    drawing.add(chart)
    return drawing


def _instrument_chart(entries: list[TradeJournal]):
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.lib import colors

    rows = instrument_pnl(entries, 5)
    drawing = Drawing(240, 140)
    drawing.add(String(8, 124, "Instrument PnL", fontSize=9, fillColor=colors.HexColor("#111827")))
    if not rows:
        drawing.add(String(18, 62, "No instrument data", fontSize=8))
        return drawing
    chart = VerticalBarChart()
    chart.x = 25
    chart.y = 25
    chart.width = 190
    chart.height = 80
    chart.data = [[value for _, value in rows]]
    chart.categoryAxis.categoryNames = [name[:8] for name, _ in rows]
    chart.bars[0].fillColor = colors.HexColor("#F7C948")
    values = [value for _, value in rows]
    chart.valueAxis.valueMin = min(values + [0])
    chart.valueAxis.valueMax = max(values + [0]) or 1
    drawing.add(chart)
    return drawing


def _emotion_chart(entries: list[TradeJournal]):
    from reportlab.graphics.charts.piecharts import Pie
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.lib import colors

    rows = emotion_counts(entries, 5)
    drawing = Drawing(240, 140)
    drawing.add(String(8, 124, "Emotion Impact", fontSize=9, fillColor=colors.HexColor("#111827")))
    if not rows:
        drawing.add(String(18, 62, "No emotion data", fontSize=8))
        return drawing
    pie = Pie()
    pie.x = 55
    pie.y = 20
    pie.width = 95
    pie.height = 95
    pie.data = [value for _, value in rows]
    pie.labels = [name[:8] for name, _ in rows]
    palette = ["#2563EB", "#F97316", "#EF4444", "#22C55E", "#A855F7"]
    for index, color in enumerate(palette[: len(rows)]):
        pie.slices[index].fillColor = colors.HexColor(color)
    drawing.add(pie)
    return drawing


def _risk_chart(entries: list[TradeJournal]):
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.lib import colors

    rows = risk_buckets(entries)
    drawing = Drawing(240, 140)
    drawing.add(String(8, 124, "Risk Distribution", fontSize=9, fillColor=colors.HexColor("#111827")))
    if not rows:
        drawing.add(String(18, 62, "No risk data", fontSize=8))
        return drawing
    chart = VerticalBarChart()
    chart.x = 25
    chart.y = 25
    chart.width = 190
    chart.height = 80
    chart.data = [[value for _, value in rows]]
    chart.categoryAxis.categoryNames = [name for name, _ in rows]
    chart.bars[0].fillColor = colors.HexColor("#14B8A6")
    chart.valueAxis.valueMin = 0
    chart.valueAxis.valueMax = max(value for _, value in rows) or 1
    drawing.add(chart)
    return drawing


def _add_excel_charts(sheet, instrument_start: int, emotion_start: int) -> None:
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference

    if sheet.max_row >= 3:
        line = LineChart()
        line.title = "Equity Curve"
        line.y_axis.title = "Cumulative PnL"
        line.x_axis.title = "Trade"
        data = Reference(sheet, min_col=2, min_row=1, max_row=max(2, instrument_start - 2))
        line.add_data(data, titles_from_data=True)
        sheet.add_chart(line, "D2")
    if instrument_start + 1 < emotion_start:
        bar = BarChart()
        bar.title = "Instrument Net PnL"
        data = Reference(sheet, min_col=2, min_row=instrument_start, max_row=emotion_start - 2)
        cats = Reference(sheet, min_col=1, min_row=instrument_start + 1, max_row=emotion_start - 2)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        sheet.add_chart(bar, "D18")
    if emotion_start + 1 <= sheet.max_row:
        pie = PieChart()
        pie.title = "Emotion Distribution"
        data = Reference(sheet, min_col=2, min_row=emotion_start, max_row=sheet.max_row)
        labels = Reference(sheet, min_col=1, min_row=emotion_start + 1, max_row=sheet.max_row)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        sheet.add_chart(pie, "D34")
